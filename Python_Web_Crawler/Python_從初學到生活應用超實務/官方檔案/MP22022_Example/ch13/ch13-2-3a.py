from openpyxl import load_workbook

wb = load_workbook("Excel資料.xlsx")
ws2 = wb.create_sheet(title="測試值2")

ws2["A1"] = "值1"
ws2["B1"] = "值2"
ws2["C1"] = "值3"

ws2.append([4, 5, 6])
ws2.append([6, 5, 4])

wb.save("Excel資料2.xlsx")   
wb.close()


