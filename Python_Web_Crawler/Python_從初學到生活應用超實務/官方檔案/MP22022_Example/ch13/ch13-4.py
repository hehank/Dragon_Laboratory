import requests, json
import pandas as pd
import time
from openpyxl import load_workbook
from datetime import datetime

url = "https://tcgbusfs.blob.core.windows.net/blobyoubike/YouBikeTP.json"
fname = "YouBike資料.xlsx"
wb = load_workbook(fname)
ws = wb.active

for i in range(10):
    r = requests.get(url)
    r.encoding = "utf-8"
    data = json.loads(r.text)
    df = pd.DataFrame.from_dict(data["retVal"], orient='index')
    tot = df.iloc[0,2]
    print("總停車格: ", tot)
    sbi = df.iloc[0,3]
    print("車輛數: ", sbi)
    bemp = df.iloc[0,12]
    print("空位數: ", bemp)
    print("----------------------------")
    maxRow = ws.max_row + 1
    ws["A"+str(maxRow)] = str(datetime.now())
    ws["B"+str(maxRow)] = int(tot)
    ws["C"+str(maxRow)] = int(sbi)
    ws["D"+str(maxRow)] = int(bemp)
    time.sleep(120)

wb.save(fname)
wb.close()