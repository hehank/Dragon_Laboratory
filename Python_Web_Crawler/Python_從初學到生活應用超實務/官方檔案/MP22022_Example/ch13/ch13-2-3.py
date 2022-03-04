from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "測試值"

ws["A1"] = "值1"
ws["B1"] = "值2"
ws["C1"] = "值3"

ws.append([1, 2, 3])
ws.append([3, 2, 1])

wb.save("Excel資料.xlsx")   
wb.close()


