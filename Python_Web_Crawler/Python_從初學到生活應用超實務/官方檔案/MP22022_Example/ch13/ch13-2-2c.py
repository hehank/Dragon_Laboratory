from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
ws = wb['工作表1']
v  = ws["A1"].value
print("A1儲存格: ", v)
v  = ws["B2"].value
print("B2儲存格: ", v)
wb.close()

