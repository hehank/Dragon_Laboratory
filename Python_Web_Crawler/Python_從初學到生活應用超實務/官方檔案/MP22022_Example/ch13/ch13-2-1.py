from openpyxl import load_workbook

wb = load_workbook("營業額.xlsx")
print(wb.sheetnames)
ws = wb.active
print("工作表名稱: ", ws.title)
print("最小欄: ", ws.min_column)
print("最大欄: ", ws.max_column)
print("最小列: ", ws.min_row)
print("最大列: ", ws.max_row)
wb.close()


