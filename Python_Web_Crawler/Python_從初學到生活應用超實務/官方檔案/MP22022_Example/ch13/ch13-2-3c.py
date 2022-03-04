from openpyxl import load_workbook

wb = load_workbook("Excel資料3.xlsx")
ws = wb["測試值2"]
ws.merge_cells("A2:C2")
ws.merge_cells("A3:C3")
wb.save("Excel資料4.xlsx")   
wb.close()


